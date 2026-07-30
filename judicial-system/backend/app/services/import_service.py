"""services/import_service.py — Excel 导入三段拆分 + 导出/模板（对照原 persons.py:592-956）
解析(read_upload/parse_excel_rows) → 校验(validate_rows) → 落库(import_persons 单事务)。
骨架依据 04-code/import_service.py，但 IMPORT_FIELD_MAP 用原版完整 24 字段
（persons.py:666-691），并保留原版全部校验：必填/校验位/文件内去重/4 个日期字段/
枚举白名单/身份证推算性别生日（校验规则只能比原版严不能松）。
"""
import io
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional, Set, Tuple
from urllib.parse import quote

from fastapi import UploadFile
from fastapi.responses import StreamingResponse
from pydantic import ValidationError
from sqlmodel import Session, select

from app.models.person import Person
from app.schemas.common import BizError, ErrorCode
from app.schemas.import_ import ImportErrorDetail, ImportResult
from app.schemas.person import PersonCreate, PersonQueryParams
from app.services.stats_service import apply_export_filters
from app.utils.id_card import infer_from_id_card, validate_id_card

MAX_IMPORT_BYTES = 10 * 1024 * 1024      # 10MB（原有限制保留）
IMPORT_CHUNK_BYTES = 1024 * 1024         # 分块读 1MB/块，超限提前终止
MAX_IMPORT_ERRORS = 100                  # 错误明细截断上限（决议 D10-3）
MAX_EXPORT_ROWS = 10000                 # 导出行数硬上限，防 OOM/超时
ALLOWED_SUFFIXES = (".xlsx",)

IMPORT_FIELD_MAP = {                     # 表头→字段映射（原版完整 24 字段，顺序不动）
    "姓名": "name", "身份证号": "id_card", "性别": "gender", "出生日期": "birth_date",
    "户籍地址": "household_addr", "现住址": "current_addr", "联系电话": "phone",
    "原罪名": "original_crime", "原判刑期": "original_sentence", "释放日期": "release_date",
    "帮教起始日期": "edu_start_date", "帮教截止日期": "edu_end_date",
    "帮教责任人": "responsible_person", "状态": "status", "风险等级": "risk_level",
    "家属姓名": "family_name", "家属电话": "family_phone", "婚姻状况": "marital_status",
    "文化程度": "education_level", "就业情况": "employment", "身体状况": "health_status",
    "经济状况": "economic_status", "备注": "notes", "责任单位": "responsible_org",
}
IMPORT_REQUIRED_FIELDS = {"name", "id_card"}
IMPORT_ENUM_MAP = {"status": ["在帮", "已解除", "脱管", "重点关注"],
                   "risk_level": ["高", "中", "低"]}
IMPORT_DATE_FIELDS = {"birth_date", "release_date", "edu_start_date", "edu_end_date"}
_FIELD_TO_HEADER = {v: k for k, v in IMPORT_FIELD_MAP.items()}

# Pydantic 英文错误消息 -> 中文翻译（导入校验用）
_PYDANTIC_MSG_ZH = {
    "String should have at most 20 characters": "不超过20个字符",
    "String should have at most 50 characters": "不超过50个字符",
    "String should have at most 100 characters": "不超过100个字符",
    "String should have at most 200 characters": "不超过200个字符",
    "String should have at most 2000 characters": "不超过2000个字符",
    "String should match pattern": "格式不正确",
    "Input should be a valid date": "日期格式不正确，应为 YYYY-MM-DD",
    "ensure this value is greater than or equal to 1": "值必须大于等于1",
    "ensure this value is less than or equal to 365": "值必须小于等于365",
    "Value error, 身份证号必须是18位": "身份证号必须是18位",
}


def _translate_pydantic_msg(msg: str) -> str:
    """将 Pydantic 英文/技术性错误消息翻译为用户友好的中文"""
    for en, zh in _PYDANTIC_MSG_ZH.items():
        if en in msg:
            return zh
    return msg


# ---------- 第 1 段：解析（纯 IO，无业务规则） ----------

async def read_upload(file: UploadFile) -> bytes:
    """分块读+提前超限终止（修复原版先全量读内存问题）"""
    if not file.filename or not file.filename.lower().endswith(ALLOWED_SUFFIXES):
        raise BizError(ErrorCode.FILE_TYPE_UNSUPPORTED, "请上传 .xlsx 格式的 Excel 文件")
    chunks: List[bytes] = []
    size = 0
    while chunk := await file.read(IMPORT_CHUNK_BYTES):
        size += len(chunk)
        if size > MAX_IMPORT_BYTES:
            raise BizError(ErrorCode.FILE_SIZE_EXCEEDED, "文件大小不能超过 10MB")
        chunks.append(chunk)
    return b"".join(chunks)


def parse_excel_rows(content: bytes) -> List[Dict[str, Any]]:
    """文件字节 → 行 dict 列表（键为中文表头）；文件级问题抛 BizError"""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h else "" for h in next(rows_iter)]
    except Exception:
        raise BizError(ErrorCode.BAD_REQUEST_BODY, "无法解析 Excel 文件")
    known = [h for h in headers if h in IMPORT_FIELD_MAP]
    if not known:
        raise BizError(ErrorCode.BAD_REQUEST_BODY, "未找到有效的表头，请使用导入模板")
    if "姓名" not in headers or "身份证号" not in headers:
        raise BizError(ErrorCode.BAD_REQUEST_BODY, "表头缺少必填列：姓名、身份证号")
    rows: List[Dict[str, Any]] = []
    for raw in rows_iter:
        if all(v is None or str(v).strip() == "" for v in raw):
            continue                                   # 空行不计入 total_rows（同原版）
        rows.append({headers[i]: raw[i] for i in range(min(len(headers), len(raw)))})
    wb.close()
    if not rows:
        raise BizError(ErrorCode.BAD_REQUEST_BODY,
                       "Excel 文件至少需要包含表头行和一行数据")
    return rows


# ---------- 第 2 段：校验（纯函数，无 IO 无 session，可单测） ----------

def validate_rows(rows: List[Dict[str, Any]]
                  ) -> Tuple[List[Tuple[int, Dict]], List[ImportErrorDetail], int]:
    """逐行校验 → (合格行(行号,payload), 错误明细, 失败行数)；坏行进 errors 不拖崩整批"""
    valid: List[Tuple[int, Dict]] = []
    errors: List[ImportErrorDetail] = []
    failed_rows = 0
    seen: Set[str] = set()
    for idx, raw in enumerate(rows, start=2):          # Excel 行号从 2（1 是表头）
        row_errors = _validate_one_row(idx, raw, seen)
        payload: Dict[str, Any] = {}
        if not row_errors:
            payload, row_errors = _to_person_payload(idx, raw)
        if row_errors:
            errors.extend(row_errors)
            failed_rows += 1
        else:
            seen.add(payload["id_card"])
            valid.append((idx, payload))
    return valid, errors, failed_rows


def _validate_one_row(idx: int, raw: Dict, seen: Set[str]) -> List[ImportErrorDetail]:
    """必填/长度/校验位/文件内去重/日期/枚举（枚举校验原版有，保留）"""
    errors: List[ImportErrorDetail] = []
    row_data = {IMPORT_FIELD_MAP.get(h, h): str(v).strip()
                for h, v in raw.items() if v is not None and str(v).strip()}
    for required in IMPORT_REQUIRED_FIELDS:
        if not row_data.get(required):
            errors.append(ImportErrorDetail(
                row=idx, field=required,
                message=f"必填字段'{_FIELD_TO_HEADER[required]}'缺失"))
    name = row_data.get("name", "")
    if name and len(name) > 20:
        errors.append(ImportErrorDetail(row=idx, field="name",
                                        message="姓名超过20个字符"))
    id_card = row_data.get("id_card", "").upper()
    if id_card:
        if not validate_id_card(id_card):
            errors.append(ImportErrorDetail(row=idx, field="id_card",
                                            message="身份证号格式错误"))
        elif id_card in seen:
            errors.append(ImportErrorDetail(row=idx, field="id_card",
                                            message="身份证号在导入文件中重复"))
    for field in IMPORT_DATE_FIELDS:
        if row_data.get(field) and not _parse_date(row_data[field]):
            errors.append(ImportErrorDetail(row=idx, field=field,
                                            message="日期格式错误,应为 YYYY-MM-DD"))
    for field, allowed in IMPORT_ENUM_MAP.items():
        value = row_data.get(field)
        if value and value not in allowed:
            errors.append(ImportErrorDetail(
                row=idx, field=field,
                message=f"值'{value}'无效,允许的值:{'/'.join(allowed)}"))
    return errors


# ---------- 第 3 段：落库（唯一事务点） ----------

def import_persons(payload: Tuple[List[Tuple[int, Dict]], List[ImportErrorDetail], int],
                   session: Session) -> ImportResult:
    """库内查重 → 批量 add → 单事务 commit；错误明细截断 100 条"""
    valid_rows, errors, failed_rows = payload
    imported = 0
    skipped = 0
    for idx, person_data in valid_rows:
        if session.exec(select(Person.id).where(
                Person.id_card == person_data["id_card"],
                Person.is_deleted == False)).first():  # noqa: E712
            errors.append(ImportErrorDetail(row=idx, field="id_card",
                                            message="身份证号已存在"))
            skipped += 1
            continue
        try:
            from app.services.person_service import _release_soft_deleted_id_card
            _release_soft_deleted_id_card(person_data["id_card"], session)
            session.add(Person(**person_data))
            imported += 1
        except Exception as e:                       # 单行落库失败则整批回滚（单事务语义）
            session.rollback()
            raise BizError(ErrorCode.INTERNAL_ERROR, f"第 {idx} 行写入失败：{e}")
    session.commit()
    return ImportResult(total_rows=len(valid_rows) + failed_rows,
                        imported=imported, skipped=skipped,
                        errors=errors[:MAX_IMPORT_ERRORS],
                        errors_truncated=len(errors) > MAX_IMPORT_ERRORS)


# ---------- 导出 / 模板（原 export_excel / import template 平移，文件流） ----------

def build_export_response(params: PersonQueryParams, session: Session) -> StreamingResponse:
    """按 PersonQueryParams 筛选导出 xlsx（筛选构建与 list 一致，见 apply_export_filters）"""
    from openpyxl import Workbook
    stmt = select(Person).where(Person.is_deleted == False)  # noqa: E712
    stmt = apply_export_filters(stmt, params)
    sort_col = getattr(Person, params.sort_by)   # sort_by 已被 Literal 白名单约束
    stmt = stmt.order_by(sort_col.asc() if params.sort_order == "asc" else sort_col.desc())
    # 硬上限防 OOM/超时
    stmt = stmt.limit(MAX_EXPORT_ROWS)
    persons = session.exec(stmt).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "人员列表"
    ws.append(["姓名", "身份证号", "性别", "年龄", "电话", "原罪名",
               "状态", "风险等级", "责任人", "帮教截止"])
    today = date.today()
    for p in persons:
        age = None
        if p.birth_date:                         # 元组比较法，闰日安全
            age = today.year - p.birth_date.year - (
                (today.month, today.day) < (p.birth_date.month, p.birth_date.day))
        ws.append([p.name, p.id_card, p.gender, age, p.phone, p.original_crime,
                   p.status, p.risk_level, p.responsible_person,
                   str(p.edu_end_date) if p.edu_end_date else None])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = quote(f"人员列表_{today.isoformat()}.xlsx")
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"})


def build_template_response() -> StreamingResponse:
    """导入模板下载（原 persons.py:733-795 平移，24 列表头+示例行+说明行）"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "人员导入模板"
    headers = list(IMPORT_FIELD_MAP.keys())
    ws.append(headers)
    ws.append(["张三", "320102199001011234", "男", "1990-01-01",
               "江苏省南京市玄武区XX路XX号", "江苏省南京市鼓楼区YY路YY号",
               "13800138000", "盗窃罪", "有期徒刑3年",
               "2020-01-15", "2020-02-01", "2023-02-01",
               "王警官", "在帮", "低", "张父", "13900139000", "已婚",
               "高中", "务农", "健康", "一般", "无特殊情况", "XX司法所"])
    ws.append(["必填", "必填,18位身份证号", "男/女(可从身份证推算)", "YYYY-MM-DD(可从身份证推算)",
               "", "", "", "", "", "YYYY-MM-DD", "YYYY-MM-DD", "YYYY-MM-DD", "",
               "在帮/已解除/脱管/重点关注", "高/中/低",
               "", "", "", "", "", "", "", "", ""])
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row_idx in (2, 3):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max_len + 4, 30)
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center")
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = quote("人员导入模板.xlsx")
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"})


# ---------- 工具 ----------

def _parse_date(v: Any) -> Optional[date]:
    """兼容 Excel 日期单元格（datetime/date）与 YYYY-MM-DD 字符串"""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(str(v).strip(), "%Y-%m-%d").date()
    except ValueError:
        return None


def _to_person_payload(idx: int, raw: Dict) -> Tuple[Dict[str, Any], List[ImportErrorDetail]]:
    """中文行 dict → Person 字段 payload；身份证推算性别/生日（原版行为保留）；
    复用 PersonCreate 契约校验（比原版严：长度/电话格式等），失败转行错误不崩溃"""
    payload: Dict[str, Any] = {}
    for col, field in IMPORT_FIELD_MAP.items():
        v = raw.get(col)
        if v is None or str(v).strip() == "":
            continue
        if field in IMPORT_DATE_FIELDS:
            parsed = _parse_date(v)
            if parsed:
                payload[field] = parsed
        else:
            payload[field] = str(v).strip()
    payload["id_card"] = payload["id_card"].upper()
    for k, v in infer_from_id_card(payload["id_card"]).items():
        if not payload.get(k):
            payload[k] = v
    try:
        PersonCreate.model_validate(payload)
    except ValidationError as e:
        return {}, [ImportErrorDetail(
            row=idx, field=str(err["loc"][0]) if err["loc"] else None,
            message=_translate_pydantic_msg(str(err["msg"]))) for err in e.errors()]
    return payload, []
