"""人员 CRUD API"""
from fastapi import APIRouter, Depends, HTTPException, Query, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select, func, or_
from datetime import datetime, timezone, timedelta, date
from typing import Optional
from urllib.parse import quote
import math
import io

from app.core.database import get_session
from app.models.person import Person
from app.models.edit_log import EditLog
from app.models.visit import Visit
from app.schemas import (
    PersonCreate, PersonUpdate, PersonResponse,
    EditLogResponse, StatsSummary, NameCount,
    PaginatedResponse, ApiResponse, ImportResult, ImportErrorDetail,
)

router = APIRouter(prefix="/api/persons", tags=["人员管理"])


def desensitize_id_card(id_card: str) -> str:
    """身份证号脱敏：前6后4可见，中间用*代替"""
    if not id_card or len(id_card) < 10:
        return id_card
    return id_card[:6] + "*" * (len(id_card) - 10) + id_card[-4:]


@router.get("/prisons")
def list_prisons(session: Session = Depends(get_session)):
    """获取所有去重的服刑场所及人数"""
    q = select(Person.prison_place, func.count().label("cnt")).where(
        Person.is_deleted == False, Person.prison_place != None, Person.prison_place != ""
    ).group_by(Person.prison_place).order_by(func.count().desc())
    results = session.exec(q).all()
    return [{"name": r[0], "count": r[1]} for r in results]


@router.get("/prisons/{prison_name}/persons")
def list_prison_persons(
    prison_name: str,
    session: Session = Depends(get_session),
):
    """获取指定监狱的人员列表"""
    persons = session.exec(
        select(Person).where(
            Person.is_deleted == False, Person.prison_place == prison_name
        ).order_by(Person.name)
    ).all()
    return [{"id": p.id, "name": p.name, "id_card": p.id_card, "status": p.status} for p in persons]


@router.get("", response_model=PaginatedResponse)
def list_persons(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=2000),
    search: Optional[str] = None,
    status: Optional[str] = None,
    risk_level: Optional[str] = None,
    responsible_person: Optional[str] = None,
    crime_contains: Optional[str] = None,
    min_age: Optional[int] = None,
    max_age: Optional[int] = None,
    is_minor: Optional[bool] = None,
    is_xj: Optional[bool] = None,
    is_mental: Optional[bool] = None,
    prison_place: Optional[str] = None,
    village: Optional[str] = None,
    expiring_within_days: Optional[int] = None,
    reveal: bool = Query(False, description="是否显示完整身份证号"),
    sort_by: str = Query("updated_at", pattern="^(name|id_card|status|risk_level|edu_end_date|created_at|updated_at)$"),
    sort_order: str = Query("desc", pattern="^(asc|desc)$"),
    session: Session = Depends(get_session),
):
    """人员列表 - 分页、筛选、排序"""
    query = select(Person).where(Person.is_deleted == False)

    # 搜索:姓名、身份证号、电话模糊匹配
    if search:
        query = query.where(
            or_(
                Person.name.contains(search),
                Person.id_card.contains(search),
                Person.phone.contains(search),
            )
        )

    # 筛选
    if status:
        query = query.where(Person.status == status)
    if risk_level:
        query = query.where(Person.risk_level == risk_level)
    if responsible_person:
        query = query.where(Person.responsible_person == responsible_person)
    if crime_contains:
        query = query.where(Person.original_crime.contains(crime_contains))
    # 年龄筛选(基于出生日期计算)
    if min_age is not None:
        max_birth = date.today().replace(year=date.today().year - min_age)
        query = query.where(Person.birth_date != None, Person.birth_date <= max_birth)
    if max_age is not None:
        min_birth = date.today().replace(year=date.today().year - max_age - 1) + timedelta(days=1)
        query = query.where(Person.birth_date != None, Person.birth_date >= min_birth)
    if is_minor is not None:
        query = query.where(Person.is_minor == is_minor)
    if is_xj is not None:
        query = query.where(Person.is_xj == is_xj)
    if is_mental is not None:
        query = query.where(Person.is_mental == is_mental)
    if prison_place:
        query = query.where(Person.prison_place == prison_place)
    if village:
        query = query.where(Person.village == village)

    # 到期筛选
    if expiring_within_days is not None:
        today = date.today()
        deadline = today + timedelta(days=expiring_within_days)
        query = query.where(Person.status == "在帮")
        query = query.where(Person.edu_end_date != None, Person.edu_end_date >= today, Person.edu_end_date <= deadline)

    # 总数
    count_query = select(func.count()).select_from(query.subquery())
    total = session.exec(count_query).one()

    # 排序
    sort_column = getattr(Person, sort_by)
    if sort_order == "desc":
        query = query.order_by(sort_column.desc())
    else:
        query = query.order_by(sort_column.asc())

    # 分页
    offset = (page - 1) * page_size
    query = query.offset(offset).limit(page_size)
    persons = session.exec(query).all()

    # 身份证号处理（reveal=true 时显示完整）+ 最近走访信息
    items = []
    for p in persons:
        d = p.model_dump()
        if not reveal:
            d["id_card"] = desensitize_id_card(d["id_card"])
        # 查询最近一次走访
        last_visit = session.exec(
            select(Visit).where(Visit.person_id == p.id).order_by(Visit.visit_date.desc()).limit(1)
        ).first()
        if last_visit:
            d["last_visit_date"] = str(last_visit.visit_date)
            d["last_visit_method"] = last_visit.visit_method
            d["last_visitor"] = last_visit.visitor
        else:
            d["last_visit_date"] = None
            d["last_visit_method"] = None
            d["last_visitor"] = None
        items.append(d)

    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        pages=math.ceil(total / page_size) if total > 0 else 0,
    )


@router.get("/stats/summary", response_model=StatsSummary)
def stats_summary(
    session: Session = Depends(get_session),
):
    """统计汇总"""
    base = select(Person).where(Person.is_deleted == False)
    total = session.exec(select(func.count()).select_from(base.subquery())).one()

    # 按状态统计
    status_map = {}
    for s in ["在帮", "已解除", "脱管", "重点关注"]:
        q = select(func.count()).select_from(base.where(Person.status == s).subquery())
        status_map[s] = session.exec(q).one()

    # 按风险等级统计
    risk_map = {}
    for level, key in [("高", "risk_high"), ("中", "risk_medium"), ("低", "risk_low")]:
        q = select(func.count()).select_from(base.where(Person.risk_level == level).subquery())
        risk_map[key] = session.exec(q).one()

    # 监狱看守所人数
    total_prison = session.exec(
        select(func.count()).select_from(
            base.where(Person.prison_place != None, Person.prison_place != "").subquery()
        )
    ).one()

    # 重点帮教对象
    total_key_target = session.exec(
        select(func.count()).select_from(
            base.where(Person.is_key_target == True).subquery()
        )
    ).one()

    # 未成年人数
    total_minor = session.exec(
        select(func.count()).select_from(
            base.where(Person.is_minor == True).subquery()
        )
    ).one()

    # xj人数
    total_xj = session.exec(
        select(func.count()).select_from(
            base.where(Person.is_xj == True).subquery()
        )
    ).one()

    # 精神疾病人数
    total_mental = session.exec(
        select(func.count()).select_from(
            base.where(Person.is_mental == True).subquery()
        )
    ).one()

    # 村/居总数
    total_village = session.exec(
        select(func.count()).select_from(
            base.where(Person.village != None, Person.village != "").distinct().subquery()
        )
    ).one()

    # 即将到期
    today = date.today()
    deadline = today + timedelta(days=90)
    expiring_soon = session.exec(
        select(func.count()).select_from(
            base.where(
                Person.status == "在帮",
                Person.edu_end_date != None,
                Person.edu_end_date >= today,
                Person.edu_end_date <= deadline,
            ).subquery()
        )
    ).one()

    # 本月/本季度新增
    month_start = today.replace(day=1)
    monthly_new = session.exec(
        select(func.count()).select_from(
            base.where(Person.created_at >= datetime(month_start.year, month_start.month, 1)).subquery()
        )
    ).one()
    quarter = (today.month - 1) // 3
    quarter_start = today.replace(month=quarter * 3 + 1, day=1)
    quarterly_new = session.exec(
        select(func.count()).select_from(
            base.where(Person.created_at >= datetime(quarter_start.year, quarter_start.month, 1)).subquery()
        )
    ).one()

    # 责任人分布
    resp_q = select(Person.responsible_person, func.count().label("cnt")).where(
        Person.is_deleted == False, Person.responsible_person != None, Person.responsible_person != ""
    ).group_by(Person.responsible_person).order_by(func.count().desc())
    responsible_distribution = [NameCount(name=r[0], count=r[1]) for r in session.exec(resp_q).all()]

    # 监狱分布
    prison_q = select(Person.prison_place, func.count().label("cnt")).where(
        Person.is_deleted == False, Person.prison_place != None, Person.prison_place != ""
    ).group_by(Person.prison_place).order_by(func.count().desc())
    prison_distribution = [NameCount(name=r[0], count=r[1]) for r in session.exec(prison_q).all()]

    # 村/居分布
    village_q = select(Person.village, func.count().label("cnt")).where(
        Person.is_deleted == False, Person.village != None, Person.village != ""
    ).group_by(Person.village).order_by(func.count().desc())
    village_distribution = [NameCount(name=r[0], count=r[1]) for r in session.exec(village_q).all()]

    return StatsSummary(
        total=total,
        total_prison=total_prison,
        total_key_target=total_key_target,
        total_minor=total_minor,
        total_xj=total_xj,
        total_mental=total_mental,
        total_village=len(village_distribution),
        在帮=status_map["在帮"],
        已解除=status_map["已解除"],
        脱管=status_map["脱管"],
        重点关注=status_map["重点关注"],
        risk_high=risk_map["risk_high"],
        risk_medium=risk_map["risk_medium"],
        risk_low=risk_map["risk_low"],
        expiring_soon=expiring_soon,
        monthly_new=monthly_new,
        quarterly_new=quarterly_new,
        responsible_distribution=responsible_distribution,
        prison_distribution=prison_distribution,
        village_distribution=village_distribution,
    )


@router.get("/stats/trend")
def stats_trend(
    months: int = Query(6, ge=1, le=24),
    session: Session = Depends(get_session),
):
    """月度趋势统计 — 最近N个月每月新增人数"""
    today = date.today()
    result = []
    for i in range(months - 1, -1, -1):
        # 计算月份
        year = today.year
        month = today.month - i
        while month <= 0:
            month += 12
            year -= 1
        start = datetime(year, month, 1)
        if month == 12:
            end = datetime(year + 1, 1, 1)
        else:
            end = datetime(year, month + 1, 1)
        cnt = session.exec(
            select(func.count()).select_from(
                select(Person).where(
                    Person.is_deleted == False,
                    Person.created_at >= start,
                    Person.created_at < end,
                ).subquery()
            )
        ).one()
        result.append({"month": f"{year}-{month:02d}", "count": cnt})
    return result


@router.get("/{person_id}/risk-score")
def get_risk_score(
    person_id: int,
    session: Session = Depends(get_session),
):
    """风险评分（0-100）"""
    person = session.get(Person, person_id)
    if not person or person.is_deleted:
        raise HTTPException(status_code=404, detail="人员不存在")

    score = 0
    factors = []

    # 风险等级权重（40分）
    risk_weights = {"高": 40, "中": 20, "低": 10}
    rw = risk_weights.get(person.risk_level, 0)
    score += rw
    factors.append({"type": "risk_level", "score": rw, "detail": f"风险等级：{person.risk_level}"})

    # 走访超期（30分）
    last_visit = session.exec(
        select(Visit).where(Visit.person_id == person_id)
        .order_by(Visit.visit_date.desc()).limit(1)
    ).first()
    today = date.today()
    if last_visit:
        days_since = (today - last_visit.visit_date).days
    else:
        if person.edu_start_date:
            days_since = (today - person.edu_start_date).days
        else:
            days_since = (today - person.created_at.date()).days

    interval = person.visit_interval_days or 90
    if days_since > interval:
        overdue = days_since - interval
        vscore = min(30, 20 + overdue)
        score += vscore
        factors.append({"type": "visit_overdue", "score": vscore, "detail": f"走访超期{overdue}天"})
    elif last_visit is None:
        score += 40
        factors.append({"type": "visit_overdue", "score": 40, "detail": "从未走访"})

    # 到期预警（30分）
    if person.edu_end_date:
        days_remaining = (person.edu_end_date - today).days
        if days_remaining <= 0:
            score += 50
            factors.append({"type": "expired", "score": 50, "detail": f"已超期{abs(days_remaining)}天"})
        elif days_remaining <= 30:
            escore = max(5, 30 - days_remaining)
            score += escore
            factors.append({"type": "expiring", "score": escore, "detail": f"剩余{days_remaining}天"})

    score = min(score, 100)
    if score >= 60:
        level = "高风险"
    elif score >= 30:
        level = "中风险"
    else:
        level = "低风险"

    return {"person_id": person_id, "score": score, "level": level, "factors": factors}


@router.post("/batch/delete", response_model=ApiResponse)
def batch_delete(
    body: dict,
    session: Session = Depends(get_session),
):
    """批量删除（软删除）"""
    ids = body.get("ids", [])
    if not ids:
        raise HTTPException(status_code=400, detail="请选择要删除的人员")
    if len(ids) > 100:
        raise HTTPException(status_code=400, detail="单次最多删除100条")

    count = 0
    for pid in ids:
        person = session.get(Person, pid)
        if person and not person.is_deleted:
            person.is_deleted = True
            person.updated_at = datetime.now(timezone.utc)
            session.add(person)
            count += 1
    session.commit()
    return ApiResponse(message=f"成功删除{count}条记录")


@router.post("/batch/status", response_model=ApiResponse)
def batch_update_status(
    body: dict,
    session: Session = Depends(get_session),
):
    """批量修改状态"""
    ids = body.get("ids", [])
    new_status = body.get("status", "")
    if not ids:
        raise HTTPException(status_code=400, detail="请选择人员")
    if new_status not in ["在帮", "已解除", "脱管", "重点关注"]:
        raise HTTPException(status_code=400, detail="状态值无效")
    if len(ids) > 100:
        raise HTTPException(status_code=400, detail="单次最多100条")

    count = 0
    for pid in ids:
        person = session.get(Person, pid)
        if person and not person.is_deleted:
            person.status = new_status
            person.updated_at = datetime.now(timezone.utc)
            session.add(person)
            count += 1
    session.commit()
    return ApiResponse(message=f"成功修改{count}条记录状态为{new_status}")


@router.post("/batch/risk", response_model=ApiResponse)
def batch_update_risk(
    body: dict,
    session: Session = Depends(get_session),
):
    """批量修改风险等级"""
    ids = body.get("ids", [])
    new_risk = body.get("risk_level", "")
    if not ids:
        raise HTTPException(status_code=400, detail="请选择人员")
    if new_risk not in ["高", "中", "低"]:
        raise HTTPException(status_code=400, detail="风险等级值无效")
    if len(ids) > 100:
        raise HTTPException(status_code=400, detail="单次最多100条")

    count = 0
    for pid in ids:
        person = session.get(Person, pid)
        if person and not person.is_deleted:
            person.risk_level = new_risk
            person.updated_at = datetime.now(timezone.utc)
            session.add(person)
            count += 1
    session.commit()
    return ApiResponse(message=f"成功修改{count}条记录风险等级为{new_risk}")


@router.get("/reports/quarterly")
def quarterly_report(
    year: Optional[int] = None,
    quarter: Optional[int] = None,
    session: Session = Depends(get_session),
):
    """季度报表"""
    today = date.today()
    if year is None:
        year = today.year
    if quarter is None:
        quarter = (today.month - 1) // 3 + 1

    q_start_month = (quarter - 1) * 3 + 1
    q_start = date(year, q_start_month, 1)
    if quarter < 4:
        q_end = date(year, q_start_month + 3, 1) - timedelta(days=1)
    else:
        q_end = date(year, 12, 31)

    base = select(Person).where(Person.is_deleted == False)

    # 季度初在册人数(季度前创建且未删除)
    existing = session.exec(
        select(func.count()).select_from(
            base.where(Person.created_at < datetime(q_start.year, q_start.month, 1)).subquery()
        )
    ).one()

    # 本季度新增
    new_count = session.exec(
        select(func.count()).select_from(
            base.where(
                Person.created_at >= datetime(q_start.year, q_start.month, 1),
                Person.created_at <= datetime(q_end.year, q_end.month, q_end.day, 23, 59, 59),
            ).subquery()
        )
    ).one()

    # 本季度解除
    # (通过修改记录查询status从在帮变为已解除的)

    # 当前在册
    current = session.exec(
        select(func.count()).select_from(base.subquery())
    ).one()

    # 在帮人数
    active = session.exec(
        select(func.count()).select_from(base.where(Person.status == "在帮").subquery())
    ).one()

    # 走访统计
    from app.models.visit import Visit
    visit_base = select(Visit).where(
        Visit.visit_date >= q_start,
        Visit.visit_date <= q_end,
    )
    total_visits = session.exec(
        select(func.count()).select_from(visit_base.subquery())
    ).one()

    visit_by_method = {}
    for method in ["上门", "电话", "视频"]:
        cnt = session.exec(
            select(func.count()).select_from(
                visit_base.where(Visit.visit_method == method).subquery()
            )
        ).one()
        visit_by_method[method] = cnt

    abnormal_visits = session.exec(
        select(func.count()).select_from(
            visit_base.where(Visit.has_abnormal == True).subquery()
        )
    ).one()

    # 走访完成率(在帮人员中,本季度有走访记录的比例)
    visited_persons = session.exec(
        select(func.count()).select_from(
            select(Visit.person_id).where(
                Visit.visit_date >= q_start,
                Visit.visit_date <= q_end,
            ).distinct().subquery()
        )
    ).one()
    visit_rate = round(visited_persons / active * 100, 1) if active > 0 else 0

    return {
        "year": year,
        "quarter": quarter,
        "period": f"{q_start.isoformat()} ~ {q_end.isoformat()}",
        "existing_at_start": existing,
        "new_this_quarter": new_count,
        "current_total": current,
        "active_count": active,
        "visits": {
            "total": total_visits,
            "上门": visit_by_method.get("上门", 0),
            "电话": visit_by_method.get("电话", 0),
            "视频": visit_by_method.get("视频", 0),
            "abnormal": abnormal_visits,
            "completion_rate": visit_rate,
        },
    }


@router.get("/export/excel")
def export_excel(
    search: Optional[str] = None,
    status: Optional[str] = None,
    risk_level: Optional[str] = None,
    session: Session = Depends(get_session),
):
    """导出人员列表为 Excel 文件"""
    from openpyxl import Workbook

    query = select(Person).where(Person.is_deleted == False)
    if search:
        query = query.where(
            or_(
                Person.name.contains(search),
                Person.id_card.contains(search),
                Person.phone.contains(search),
            )
        )
    if status:
        query = query.where(Person.status == status)
    if risk_level:
        query = query.where(Person.risk_level == risk_level)

    query = query.order_by(Person.updated_at.desc())
    persons = session.exec(query).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "人员列表"

    # 表头
    headers = ["姓名", "身份证号", "性别", "年龄", "电话", "原罪名", "状态", "风险等级", "责任人", "帮教截止"]
    ws.append(headers)

    # 数据行
    today = date.today()
    for p in persons:
        age = None
        if p.birth_date:
            age = today.year - p.birth_date.year - (
                (today.month, today.day) < (p.birth_date.month, p.birth_date.day)
            )
        ws.append([
            p.name,
            p.id_card,
            p.gender,
            age,
            p.phone,
            p.original_crime,
            p.status,
            p.risk_level,
            p.responsible_person,
            str(p.edu_end_date) if p.edu_end_date else None,
        ])

    # 写入内存流
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"人员列表_{today.isoformat()}.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )


# ========== Excel 批量导入 ==========

# 字段映射:中文表头 → 字段名
IMPORT_FIELD_MAP = {
    "姓名": "name",
    "身份证号": "id_card",
    "性别": "gender",
    "出生日期": "birth_date",
    "户籍地址": "household_addr",
    "现住址": "current_addr",
    "联系电话": "phone",
    "原罪名": "original_crime",
    "原判刑期": "original_sentence",
    "释放日期": "release_date",
    "帮教起始日期": "edu_start_date",
    "帮教截止日期": "edu_end_date",
    "帮教责任人": "responsible_person",
    "状态": "status",
    "风险等级": "risk_level",
    "家属姓名": "family_name",
    "家属电话": "family_phone",
    "婚姻状况": "marital_status",
    "文化程度": "education_level",
    "就业情况": "employment",
    "身体状况": "health_status",
    "经济状况": "economic_status",
    "备注": "notes",
    "责任单位": "responsible_org",
}

IMPORT_REQUIRED_FIELDS = {"name", "id_card"}
IMPORT_ENUM_MAP = {
    "status": ["在帮", "已解除", "脱管", "重点关注"],
    "risk_level": ["高", "中", "低"],
}
IMPORT_DATE_FIELDS = {"birth_date", "release_date", "edu_start_date", "edu_end_date"}

# 身份证校验位算法
ID_CARD_WEIGHTS = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2]
ID_CARD_CHECK_CHARS = "10X98765432"


def _validate_id_card(id_card: str) -> bool:
    """校验18位身份证号(含校验位)"""
    if not id_card or len(id_card) != 18:
        return False
    body = id_card[:17]
    if not body.isdigit():
        return False
    total = sum(int(body[i]) * ID_CARD_WEIGHTS[i] for i in range(17))
    expected = ID_CARD_CHECK_CHARS[total % 11]
    return id_card[17].upper() == expected


def _infer_from_id_card(id_card: str) -> dict:
    """从身份证号推算性别和出生日期"""
    result = {}
    if id_card and len(id_card) == 18 and id_card[:17].isdigit():
        # 性别:第17位奇数为男,偶数为女
        result["gender"] = "男" if int(id_card[16]) % 2 == 1 else "女"
        # 出生日期
        try:
            result["birth_date"] = datetime.strptime(id_card[6:14], "%Y%m%d").date()
        except ValueError:
            pass
    return result


@router.get("/import/template")
def download_import_template(
):
    """下载 Excel 导入模板"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "人员导入模板"

    # 表头行
    headers = list(IMPORT_FIELD_MAP.keys())
    ws.append(headers)

    # 示例数据行
    ws.append([
        "张三", "320102199001011234", "男", "1990-01-01",
        "江苏省南京市玄武区XX路XX号", "江苏省南京市鼓楼区YY路YY号",
        "13800138000", "盗窃罪", "有期徒刑3年",
        "2020-01-15", "2020-02-01", "2023-02-01",
        "王警官", "在帮", "低",
        "张父", "13900139000", "已婚",
        "高中", "务农", "健康",
        "一般", "无特殊情况", "XX司法所",
    ])

    # 字段说明行
    ws.append([
        "必填", "必填,18位身份证号", "男/女(可从身份证推算)", "YYYY-MM-DD(可从身份证推算)",
        "", "", "",
        "", "", "YYYY-MM-DD",
        "YYYY-MM-DD", "YYYY-MM-DD", "",
        "在帮/已解除/脱管/重点关注", "高/中/低",
        "", "", "",
        "", "", "", "", "", "",
    ])

    # 列宽自适应
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        # 检查第2、3行
        for row_idx in [2, 3]:
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max_len + 4, 30)

    # 表头样式
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center")

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = quote("人员导入模板.xlsx")
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.post("/import/excel", response_model=ImportResult)
def import_excel(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
):
    """Excel 批量导入人员"""
    from openpyxl import load_workbook

    # 验证文件类型
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 格式的 Excel 文件")

    # 流式读取文件内容
    content = file.file.read()
    if len(content) > 10 * 1024 * 1024:  # 10MB 限制
        raise HTTPException(status_code=400, detail="文件大小不能超过 10MB")

    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="无法解析 Excel 文件")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel 文件至少需要包含表头行和一行数据")

    # 解析表头映射
    header_row = rows[0]
    col_map = {}  # column_index -> field_name
    for col_idx, cell_val in enumerate(header_row):
        if cell_val and str(cell_val).strip() in IMPORT_FIELD_MAP:
            col_map[col_idx] = IMPORT_FIELD_MAP[str(cell_val).strip()]

    if not col_map:
        raise HTTPException(status_code=400, detail="未找到有效的表头,请使用导入模板")

    # 预查询已存在的身份证号
    existing_id_cards = set()
    result = ImportResult(success=True, total_rows=0, imported=0, skipped=0, errors=[])

    # 收集需要批量插入的数据
    persons_to_add = []

    for row_idx, row in enumerate(rows[1:], start=2):  # 从第2行开始(第1行是表头)
        result.total_rows += 1

        # 跳过完全空行
        if not any(row):
            result.total_rows -= 1
            continue

        # 提取字段值
        row_data = {}
        for col_idx, field_name in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                val_str = str(val).strip()
                if val_str:
                    row_data[field_name] = val_str

        # 校验必填字段
        row_has_error = False
        for required_field in IMPORT_REQUIRED_FIELDS:
            if required_field not in row_data or not row_data[required_field]:
                # 找到对应的中文名
                cn_name = next((k for k, v in IMPORT_FIELD_MAP.items() if v == required_field), required_field)
                result.errors.append(ImportErrorDetail(
                    row=row_idx, field=required_field, message=f"必填字段'{cn_name}'缺失"
                ))
                row_has_error = True

        if row_has_error:
            result.skipped += 1
            continue

        # 身份证号格式校验
        id_card = row_data.get("id_card", "").upper()
        if not _validate_id_card(id_card):
            result.errors.append(ImportErrorDetail(
                row=row_idx, field="id_card", message="身份证号格式错误"
            ))
            result.skipped += 1
            continue

        row_data["id_card"] = id_card

        # 身份证号唯一性检查(内存 + 数据库)
        if id_card in existing_id_cards:
            result.errors.append(ImportErrorDetail(
                row=row_idx, field="id_card", message="身份证号在导入文件中重复"
            ))
            result.skipped += 1
            continue

        db_exists = session.exec(
            select(Person.id).where(Person.id_card == id_card, Person.is_deleted == False)
        ).first()
        if db_exists:
            result.errors.append(ImportErrorDetail(
                row=row_idx, field="id_card", message="身份证号已存在"
            ))
            result.skipped += 1
            continue

        existing_id_cards.add(id_card)

        # 日期格式校验
        for date_field in IMPORT_DATE_FIELDS:
            if date_field in row_data:
                try:
                    row_data[date_field] = datetime.strptime(row_data[date_field], "%Y-%m-%d").date()
                except ValueError:
                    result.errors.append(ImportErrorDetail(
                        row=row_idx, field=date_field, message=f"日期格式错误,应为 YYYY-MM-DD"
                    ))
                    row_has_error = True

        if row_has_error:
            result.skipped += 1
            continue

        # 枚举值校验
        for field_name, allowed_values in IMPORT_ENUM_MAP.items():
            if field_name in row_data and row_data[field_name] not in allowed_values:
                result.errors.append(ImportErrorDetail(
                    row=row_idx, field=field_name,
                    message=f"值'{row_data[field_name]}'无效,允许的值:{'/'.join(allowed_values)}"
                ))
                row_has_error = True

        if row_has_error:
            result.skipped += 1
            continue

        # 从身份证自动推算性别和出生日期(如果未提供)
        inferred = _infer_from_id_card(id_card)
        for k, v in inferred.items():
            if k not in row_data or not row_data[k]:
                row_data[k] = v

        persons_to_add.append(row_data)

    # 批量插入
    for person_data in persons_to_add:
        person = Person(**person_data)
        session.add(person)
        result.imported += 1

    if result.imported > 0:
        session.commit()

    if result.skipped > 0:
        result.success = result.imported > 0

    return result


@router.get("/{person_id}", response_model=PersonResponse)
def get_person(
    person_id: int,
    session: Session = Depends(get_session),
):
    """获取单个人员详情"""
    person = session.get(Person, person_id)
    if not person or person.is_deleted:
        raise HTTPException(status_code=404, detail="人员不存在")
    return person


@router.get("/{person_id}/edit-logs", response_model=list[EditLogResponse])
def get_edit_logs(
    person_id: int,
    session: Session = Depends(get_session),
):
    """获取人员修改历史 - 按时间倒序"""
    # 确保人员存在
    person = session.get(Person, person_id)
    if not person or person.is_deleted:
        raise HTTPException(status_code=404, detail="人员不存在")

    logs = session.exec(
        select(EditLog)
        .where(EditLog.table_name == "persons", EditLog.record_id == person_id)
        .order_by(EditLog.edited_at.desc())
    ).all()
    return logs


@router.post("", response_model=PersonResponse, status_code=201)
def create_person(
    data: PersonCreate,
    session: Session = Depends(get_session),
):
    """新增人员"""
    # 身份证号唯一检查
    existing = session.exec(
        select(Person).where(Person.id_card == data.id_card, Person.is_deleted == False)
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="该身份证号已存在")

    # 从身份证推算性别和出生日期
    person_data = data.model_dump()
    if data.id_card and len(data.id_card) == 18:
        if not person_data.get("gender"):
            person_data["gender"] = "男" if int(data.id_card[16]) % 2 == 1 else "女"
        if not person_data.get("birth_date"):
            try:
                person_data["birth_date"] = datetime.strptime(
                    data.id_card[6:14], "%Y%m%d"
                ).date()
            except ValueError:
                pass

    # 根据风险等级设置默认走访间隔(仅当用户未显式设置时)
    risk_interval_map = {"高": 30, "中": 90, "低": 180}
    if data.visit_interval_days is None or data.visit_interval_days == 90:  # 90是schema默认值
        risk = person_data.get("risk_level", "低")
        person_data["visit_interval_days"] = risk_interval_map.get(risk, 90)

    person = Person(**person_data)
    session.add(person)
    session.commit()
    session.refresh(person)
    return person


@router.put("/{person_id}", response_model=PersonResponse)
def update_person(
    person_id: int,
    data: PersonUpdate,
    session: Session = Depends(get_session),
):
    """修改人员信息 - 自动留痕"""
    person = session.get(Person, person_id)
    if not person or person.is_deleted:
        raise HTTPException(status_code=404, detail="人员不存在")

    update_data = data.model_dump(exclude_unset=True)
    editor_name = update_data.pop("editor", None) or "系统"

    # 记录变更
    for field, new_value in update_data.items():
        old_value = getattr(person, field)
        if str(old_value) != str(new_value):
            log = EditLog(
                table_name="persons",
                record_id=person_id,
                field_name=field,
                old_value=str(old_value) if old_value is not None else None,
                new_value=str(new_value) if new_value is not None else None,
                editor=editor_name,
            )
            session.add(log)

    # 更新字段
    for field, value in update_data.items():
        setattr(person, field, value)

    person.updated_at = datetime.now(timezone.utc)
    session.add(person)
    session.commit()
    session.refresh(person)
    return person


@router.delete("/{person_id}", response_model=ApiResponse)
def delete_person(
    person_id: int,
    session: Session = Depends(get_session),
):
    """删除人员(软删除)"""
    person = session.get(Person, person_id)
    if not person or person.is_deleted:
        raise HTTPException(status_code=404, detail="人员不存在")

    person.is_deleted = True
    person.updated_at = datetime.now(timezone.utc)
    session.add(person)
    session.commit()

    return ApiResponse(message="删除成功")
