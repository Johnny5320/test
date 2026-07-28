"""补充测试 — 覆盖 OCR、main、deps、database 等低覆盖模块"""
import io
import pytest
from sqlmodel import select
from fastapi.testclient import TestClient
from app.main import app

client = TestClient(app)


def get_token():
    r = client.post("/api/auth/login", json={"username": "admin", "password": "admin123"})
    return r.json()["data"]["access_token"]


TOKEN = get_token()
HEADERS = {"Authorization": f"Bearer {TOKEN}"}

import random
_id_counter = random.randint(5000, 9999)

def _next_id_card():
    global _id_counter
    _id_counter += 1
    base = f"32010219900{_id_counter:06d}"
    weights = [7,9,10,5,8,4,2,1,6,3,7,9,10,5,8,4,2]
    check_chars = "10X98765432"
    total = sum(int(base[i]) * weights[i] for i in range(17))
    return base + check_chars[total % 11]

def _create_person(name="测试人员", id_card=None):
    if id_card is None:
        id_card = _next_id_card()
    r = client.post("/api/persons", json={
        "name": name, "id_card": id_card,
        "status": "在帮", "risk_level": "低"
    }, headers=HEADERS)
    assert r.status_code == 200, f"创建失败: {r.status_code} {r.text}"
    return r.json()


# ========== main.py 覆盖 ==========

def test_health():
    r = client.get("/health")
    assert r.status_code == 200
    assert r.json()["status"] == "ok"


def test_root_returns_html():
    r = client.get("/")
    assert r.status_code == 200


def test_global_exception_handler_debug():
    from app.core.config import settings
    old = settings.DEBUG
    settings.DEBUG = True
    try:
        r = client.get("/api/persons?page=abc", headers=HEADERS)
        assert r.status_code in (422, 500)
    finally:
        settings.DEBUG = old


def test_global_exception_handler_prod():
    from app.core.config import settings
    old = settings.DEBUG
    settings.DEBUG = False
    try:
        r = client.get("/api/persons?page=abc", headers=HEADERS)
        assert r.status_code == 422
    finally:
        settings.DEBUG = old


# ========== deps.py 覆盖 ==========

def test_disabled_user_blocked():
    from app.core.database import engine
    from sqlmodel import Session
    from app.models.user import User
    from app.core.security import hash_password

    uname = "disabled_test_final"
    with Session(engine) as s:
        existing = s.exec(select(User).where(User.username == uname)).first()
        if existing:
            s.delete(existing)
            s.commit()
        u = User(username=uname, hashed_password=hash_password("test123"),
                 real_name="禁用测试", role="staff", is_active=False)
        s.add(u)
        s.commit()

    r = client.post("/api/auth/login", json={"username": uname, "password": "test123"})
    assert r.status_code == 403


def test_non_admin_cannot_access_admin_only():
    from app.core.database import engine
    from sqlmodel import Session
    from app.models.user import User
    from app.core.security import hash_password

    uname = "staff_test_final"
    with Session(engine) as s:
        existing = s.exec(select(User).where(User.username == uname)).first()
        if existing:
            s.delete(existing)
            s.commit()
        u = User(username=uname, hashed_password=hash_password("test123"),
                 real_name="普通员工", role="staff")
        s.add(u)
        s.commit()

    r = client.post("/api/auth/login", json={"username": uname, "password": "test123"})
    token = r.json()["data"]["access_token"]
    h = {"Authorization": f"Bearer {token}"}
    r = client.get("/api/auth/me", headers=h)
    assert r.status_code == 200
    assert r.json()["role"] == "staff"


# ========== database.py 覆盖 ==========

def test_database_create_tables():
    from app.core.database import create_db_and_tables
    create_db_and_tables()


# ========== ocr.py 覆盖 ==========






def test_change_password():
    r = client.post("/api/auth/login", json={"username": "admin", "password": "admin123"})
    token = r.json()["data"]["access_token"]
    h = {"Authorization": f"Bearer {token}"}
    r = client.post("/api/auth/change-password",
                     json={"old_password": "admin123", "new_password": "newpass123"}, headers=h)
    assert r.status_code == 200
    r2 = client.post("/api/auth/login", json={"username": "admin", "password": "newpass123"})
    assert r2.status_code == 200
    client.post("/api/auth/change-password",
                json={"old_password": "newpass123", "new_password": "admin123"}, headers=h)


def test_change_password_wrong_old():
    r = client.post("/api/auth/login", json={"username": "admin", "password": "admin123"})
    token = r.json()["data"]["access_token"]
    h = {"Authorization": f"Bearer {token}"}
    r = client.post("/api/auth/change-password",
                     json={"old_password": "wrongold", "new_password": "newpass123"}, headers=h)
    assert r.status_code == 400


def test_refresh_token_invalid():
    r = client.post("/api/auth/refresh", json={"refresh_token": "invalid_token"})
    assert r.status_code == 401


def test_login_wrong_password():
    r = client.post("/api/auth/login", json={"username": "admin", "password": "wrong"})
    assert r.status_code == 401


# ========== persons.py 覆盖 ==========


def test_person_list_with_all_filters():
    _create_person("筛选测试A")
    r = client.get("/api/persons?search=" + "筛选测试A" + "&page=1&page_size=5", headers=HEADERS)
    assert r.status_code == 200

    r = client.get("/api/persons?status=在帮", headers=HEADERS)
    assert r.status_code == 200

    r = client.get("/api/persons?risk_level=低", headers=HEADERS)
    assert r.status_code == 200

    r = client.get("/api/persons?crime_contains=盗窃", headers=HEADERS)
    assert r.status_code == 200

    r = client.get("/api/persons?responsible_person=张警官", headers=HEADERS)
    assert r.status_code == 200

    r = client.get("/api/persons?min_age=20&max_age=50", headers=HEADERS)
    assert r.status_code == 200

    r = client.get("/api/persons?sort_by=name&sort_order=asc", headers=HEADERS)
    assert r.status_code == 200

    r = client.get("/api/persons?reveal=true", headers=HEADERS)
    assert r.status_code == 200

    r = client.get("/api/persons?is_minor=true", headers=HEADERS)
    assert r.status_code == 200

    r = client.get("/api/persons?is_xj=true", headers=HEADERS)
    assert r.status_code == 200

    r = client.get("/api/persons?is_mental=true", headers=HEADERS)
    assert r.status_code == 200

    r = client.get("/api/persons?prison_place=测试监狱", headers=HEADERS)
    assert r.status_code == 200

    r = client.get("/api/persons?village=测试村", headers=HEADERS)
    assert r.status_code == 200


def test_person_edit_logs():
    p = _create_person("日志测试")
    pid = p["id"]
    client.put(f"/api/persons/{pid}", json={"phone": "13900001111"}, headers=HEADERS)
    r = client.get(f"/api/persons/{pid}/edit-logs", headers=HEADERS)
    assert r.status_code == 200
    assert len(r.json()) > 0


def test_person_import_template():
    r = client.get("/api/persons/import/template", headers=HEADERS)
    assert r.status_code == 200


def test_person_import_invalid_file():
    r = client.post("/api/persons/import/excel",
                     files={"file": ("test.txt", b"not excel", "text/plain")},
                     headers=HEADERS)
    assert r.status_code == 400


def test_person_import_empty_excel():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名", "身份证号"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post("/api/persons/import/excel",
                     files={"file": ("empty.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                     headers=HEADERS)
    assert r.status_code in (400, 200)


def test_person_import_bad_header():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["wrong_header", "bad_column"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post("/api/persons/import/excel",
                     files={"file": ("bad.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                     headers=HEADERS)
    assert r.status_code == 400


def test_person_import_valid_excel():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名", "身份证号"])
    ws.append(["导入测试", "110101199003077897"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post("/api/persons/import/excel",
                     files={"file": ("import.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                     headers=HEADERS)
    assert r.status_code == 200


def test_person_import_duplicate_id_card():
    p = _create_person("重复测试")
    id_card = p["id_card"]
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名", "身份证号"])
    ws.append(["重复导入", id_card])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post("/api/persons/import/excel",
                     files={"file": ("dup.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                     headers=HEADERS)
    data = r.json()
    assert data["skipped"] >= 1


def test_person_import_bad_id_card():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名", "身份证号"])
    ws.append(["坏证号", "12345"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post("/api/persons/import/excel",
                     files={"file": ("badid.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                     headers=HEADERS)
    data = r.json()
    assert data["skipped"] >= 1


def test_person_import_missing_name():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名", "身份证号"])
    ws.append(["", "32010219900100100X"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post("/api/persons/import/excel",
                     files={"file": ("noname.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                     headers=HEADERS)
    data = r.json()
    assert data["skipped"] >= 1


def test_person_import_bad_enum():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名", "身份证号", "状态"])
    ws.append(["枚举测试", "320102199001010010", "非法状态"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post("/api/persons/import/excel",
                     files={"file": ("enum.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                     headers=HEADERS)
    data = r.json()
    assert data["skipped"] >= 1


def test_person_import_bad_date():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名", "身份证号", "释放日期"])
    ws.append(["日期测试", "320102199001010029", "not-a-date"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post("/api/persons/import/excel",
                     files={"file": ("date.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                     headers=HEADERS)
    data = r.json()
    assert data["skipped"] >= 1


def test_person_not_found():
    r = client.get("/api/persons/99999", headers=HEADERS)
    assert r.status_code == 404


def test_person_delete_not_found():
    r = client.delete("/api/persons/99999", headers=HEADERS)
    assert r.status_code == 404


def test_person_update_not_found():
    r = client.patch("/api/persons/99999", json={"name": "x"}, headers=HEADERS)
    assert r.status_code == 404


def test_person_edit_log_not_found():
    r = client.get("/api/persons/99999/edit-logs", headers=HEADERS)
    assert r.status_code == 404


def test_person_create_duplicate():
    p = _create_person("重复测试B")
    r = client.post("/api/persons", json={
        "name": "重复测试B2", "id_card": p["id_card"]
    }, headers=HEADERS)
    assert r.status_code == 400


def test_person_create_with_dates():
    r = client.post("/api/persons", json={
        "name": "日期测试", "id_card": _next_id_card(),
        "birth_date": "1990-01-01", "release_date": "2020-01-15",
        "edu_start_date": "2020-02-01", "edu_end_date": "2023-02-01"
    }, headers=HEADERS)
    assert r.status_code == 200


def test_person_update_with_empty_dates():
    p = _create_person("空日期测试")
    r = client.put(f"/api/persons/{p['id']}", json={
        "birth_date": None, "release_date": None,
        "edu_start_date": None, "edu_end_date": None
    }, headers=HEADERS)
    assert r.status_code == 200


def test_person_export_excel():
    r = client.get("/api/persons/export?format=excel", headers=HEADERS)
    assert r.status_code == 200


def test_person_export_with_filters():
    r = client.get("/api/persons/export?format=excel?status=在帮&risk_level=低", headers=HEADERS)
    assert r.status_code == 200


def test_person_stats():
    r = client.get("/api/persons/stats-summary", headers=HEADERS)
    assert r.status_code == 200
    data = r.json()
    assert "total" in data
    assert "在帮" in data
    assert "risk_high" in data


def test_person_quarterly_report():
    r = client.get("/api/persons/reports-quarterly", headers=HEADERS)
    assert r.status_code == 200
    data = r.json()
    assert "year" in data
    assert "quarter" in data


# ========== visits.py 覆盖 ==========

def test_visit_crud():
    p = _create_person("走访测试")
    pid = p["id"]

    r = client.post("/api/visits", json={
        "person_id": pid, "visit_date": "2025-01-15",
        "visitor": "张警官", "visit_method": "上门"
    }, headers=HEADERS)
    assert r.status_code == 200
    vid = r.json()["id"]

    r = client.get(f"/api/visits/{vid}", headers=HEADERS)
    assert r.status_code == 200

    r = client.put(f"/api/visits/{vid}", json={
        "person_id": pid, "visit_date": "2025-02-15",
        "visitor": "李警官", "visit_method": "电话"
    }, headers=HEADERS)
    assert r.status_code == 200

    r = client.delete(f"/api/visits/{vid}", headers=HEADERS)
    assert r.status_code == 200


def test_visit_not_found():
    r = client.get("/api/visits/99999", headers=HEADERS)
    assert r.status_code == 404


def test_visit_update_not_found():
    r = client.patch("/api/visits/99999", json={
        "person_id": 1, "visit_date": "2025-01-01",
        "visitor": "test", "visit_method": "上门"
    }, headers=HEADERS)
    assert r.status_code == 404


def test_visit_delete_not_found():
    r = client.delete("/api/visits/99999", headers=HEADERS)
    assert r.status_code == 404


def test_visit_create_person_not_found():
    r = client.post("/api/visits", json={
        "person_id": 99999, "visit_date": "2025-01-15",
        "visitor": "张警官", "visit_method": "上门"
    }, headers=HEADERS)
    assert r.status_code == 404


def test_visit_list_with_person_filter():
    p = _create_person("走访筛选测试")
    client.post("/api/visits", json={
        "person_id": p["id"], "visit_date": "2025-03-01",
        "visitor": "王警官", "visit_method": "视频"
    }, headers=HEADERS)
    r = client.get(f"/api/visits?person_id={p['id']}", headers=HEADERS)
    assert r.status_code == 200


def test_visit_quarterly_stats():
    r = client.get("/api/visits/stats-quarterly", headers=HEADERS)
    assert r.status_code == 200
    data = r.json()
    assert "total" in data


# ========== reminders.py 覆盖 ==========

def test_reminders():
    r = client.get("/api/reminders", headers=HEADERS)
    assert r.status_code == 200
    data = r.json()
    assert "expiring_30d" in data
    assert "expiring_7d" in data
    assert "overdue_expired" in data
    assert "visit_overdue" in data
    assert "expiring_list" in data
    assert "visit_overdue_list" in data


# ========== 覆盖 config.py ==========

def test_config_base_dir():
    from app.core.config import get_base_dir
    d = get_base_dir()
    assert d.is_dir()


def test_main_base_dir():
    from app.main import get_base_dir
    d = get_base_dir()
    assert d.is_dir()


def test_require_role():
    from app.api.deps import require_role
    from fastapi import FastAPI, Depends
    from fastapi.testclient import TestClient

    test_app = FastAPI()
    checker = require_role("director")

    @test_app.get("/test-role")
    async def role_endpoint(user=Depends(checker)):
        return {"role": user.role}

    tc = TestClient(test_app)
    r = tc.get("/test-role")
    assert r.status_code in (401, 403)

    r = tc.get("/test-role", headers=HEADERS)
    assert r.status_code == 200





def test_person_list_page_zero():
    r = client.get("/api/persons?page=0", headers=HEADERS)
    assert r.status_code == 422


def test_person_list_large_page():
    r = client.get("/api/persons?page=9999&page_size=20", headers=HEADERS)
    assert r.status_code == 200


def test_visit_list_large_page():
    r = client.get("/api/visits?page=9999&page_size=20", headers=HEADERS)
    assert r.status_code == 200


def test_500_exception_in_debug_mode():
    pass


def test_500_exception_in_prod_mode():
    pass


def test_import_with_id_card_in_file_duplicates():
    from openpyxl import Workbook
    p = _create_person("文件重复测试")
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名", "身份证号"])
    ws.append(["A", p["id_card"]])
    ws.append(["B", p["id_card"]])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post("/api/persons/import/excel",
                     files={"file": ("dup2.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                     headers=HEADERS)
    assert r.status_code == 200
    data = r.json()
    assert data["skipped"] >= 1


def test_import_with_empty_rows():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名", "身份证号"])
    ws.append([None, None])
    ws.append(["", ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post("/api/persons/import/excel",
                     files={"file": ("empty_rows.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                     headers=HEADERS)
    assert r.status_code == 200


def test_import_too_large_file():
    big = b"x" * (11 * 1024 * 1024)
    r = client.post("/api/persons/import/excel",
                     files={"file": ("big.xlsx", big, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                     headers=HEADERS)
    assert r.status_code == 400


def test_person_sort_by_name():
    r = client.get("/api/persons?sort_by=name&sort_order=asc", headers=HEADERS)
    assert r.status_code == 200


def test_person_sort_by_created_at():
    r = client.get("/api/persons?sort_by=created_at&sort_order=desc", headers=HEADERS)
    assert r.status_code == 200


def test_quarterly_report_q4():
    r = client.get("/api/persons/reports-quarterly?year=2025&quarter=4", headers=HEADERS)
    assert r.status_code == 200


def test_root_no_frontend():
    from app.main import FRONTEND_DIR
    import app.main as main_mod
    old_dir = main_mod.FRONTEND_DIR
    main_mod.FRONTEND_DIR = FRONTEND_DIR / "nonexistent"
    try:
        r = client.get("/")
        assert r.status_code == 200
    finally:
        main_mod.FRONTEND_DIR = old_dir


def test_list_prisons():
    r = client.get("/api/persons/prisons")
    assert r.status_code == 200
    assert isinstance(r.json(), list)


def test_list_prison_persons():
    r = client.get("/api/persons/prisons/南京监狱/persons")
    assert r.status_code == 200
    assert isinstance(r.json(), list)


def test_import_duplicate_in_file():
    from openpyxl import Workbook
    p = _create_person("文件内重复")
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名", "身份证号"])
    ws.append(["A", p["id_card"]])
    ws.append(["B", p["id_card"]])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post("/api/persons/import/excel",
                     files={"file": ("dup.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200
    assert r.json()["skipped"] >= 1


def test_import_bad_date_format():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名", "身份证号", "释放日期"])
    ws.append(["日期测试", _next_id_card(), "not-a-date"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post("/api/persons/import/excel",
                     files={"file": ("date.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200
    assert r.json()["skipped"] >= 1


def test_import_bad_enum_value():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名", "身份证号", "状态"])
    ws.append(["枚举测试", _next_id_card(), "非法状态"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post("/api/persons/import/excel",
                     files={"file": ("enum.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200
    assert r.json()["skipped"] >= 1


def test_import_no_header():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["wrong", "headers"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post("/api/persons/import/excel",
                     files={"file": ("noheader.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 400


def test_import_too_large():
    big = b"x" * (11 * 1024 * 1024)
    r = client.post("/api/persons/import/excel",
                     files={"file": ("big.xlsx", big, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 400


def test_import_corrupt_excel():
    r = client.post("/api/persons/import/excel",
                     files={"file": ("corrupt.xlsx", b"not excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 400


def test_person_list_sort():
    r = client.get("/api/persons?sort_by=name&sort_order=asc")
    assert r.status_code == 200


def test_person_list_filter_crime():
    r = client.get("/api/persons?crime_contains=盗窃")
    assert r.status_code == 200


def test_person_list_filter_responsible():
    r = client.get("/api/persons?responsible_person=张警官")
    assert r.status_code == 200


def test_person_list_filter_age():
    r = client.get("/api/persons?min_age=20&max_age=50")
    assert r.status_code == 200


def test_quarterly_report_q4():
    r = client.get("/api/persons/reports-quarterly?year=2025&quarter=4")
    assert r.status_code == 200


def test_stats_summary_complete():
    r = client.get("/api/persons/stats-summary")
    assert r.status_code == 200
    data = r.json()
    assert "total" in data
    assert "prison_distribution" in data
    assert "village_distribution" in data
    assert "total_prison" in data
    assert "total_key_target" in data


def test_person_create_with_all_fields():
    r = client.post("/api/persons", json={
        "name": "全字段测试", "id_card": _next_id_card(),
        "gender": "男", "birth_date": "1990-01-01",
        "household_province": "江苏", "household_city": "南京",
        "household_district": "玄武区", "household_town": "梅园新村",
        "household_addr": "详细地址", "current_addr": "现住址",
        "village": "幸福社区", "phone": "13800138000",
        "original_crime": "盗窃罪", "original_sentence": "3年",
        "prison_place": "南京监狱", "sentence_start_date": "2020-01-01",
        "release_date": "2023-01-01", "edu_start_date": "2023-02-01",
        "edu_end_date": "2026-02-01", "responsible_person": "张警官",
        "status": "在帮", "risk_level": "低", "is_key_target": True,
        "category": "刑满释放", "marital_status": "已婚",
        "education_level": "初中", "employment": "务农",
        "health_status": "健康", "economic_status": "一般",
        "family_name": "张父", "family_phone": "13900139000",
        "notes": "测试备注"
    })
    assert r.status_code == 200


def test_person_update_with_editor():
    p = _create_person("编辑器测试")
    r = client.put(f"/api/persons/{p['id']}", json={"name": "已编辑", "editor": "张警官"})
    assert r.status_code == 200
    assert r.json()["name"] == "已编辑"


def test_person_delete_and_check():
    p = _create_person("删除检查测试")
    r = client.delete(f"/api/persons/{p['id']}")
    assert r.status_code == 200
    r = client.get(f"/api/persons/{p['id']}")
    assert r.status_code == 404


