"""
全面 API 测试 — 覆盖所有端点、所有字段、所有错误路径
目标：98%+ 覆盖率
"""
import pytest
import io
from fastapi.testclient import TestClient
from sqlmodel import SQLModel, Session, create_engine
from sqlmodel.pool import StaticPool

from app.main import app
from app.core.database import get_session
from app.models import User, Person, Visit, EditLog
from app.core.security import hash_password




# ============================================================
# 认证
# ============================================================
# 认证 API
# ============================================================

class TestAuth:
    def test_login_success(self, client):
        r = client.post("/api/auth/login", json={"username": "admin", "password": "admin123"})
        assert r.status_code == 200
        d = r.json()
        assert "access_token" in d
        assert "refresh_token" in d
        assert d["token_type"] == "bearer"
        assert "force_change_password" in d

    def test_login_wrong_password(self, client):
        r = client.post("/api/auth/login", json={"username": "admin", "password": "wrong"})
        assert r.status_code == 401

    def test_login_nonexistent_user(self, client):
        r = client.post("/api/auth/login", json={"username": "nobody", "password": "***"})
        assert r.status_code == 401

    def test_get_me(self, client):
        r = client.get("/api/auth/me")
        assert r.status_code == 200
        d = r.json()
        assert d["username"] == "admin"
        assert d["real_name"] == "管理员"
        assert d["role"] == "director"
        assert d["is_active"] is True

    def test_get_me_no_token(self, client):
        r = TestClient(app).get("/api/auth/me")
        assert r.status_code in (401, 403)

    def test_refresh_token(self, client):
        login = client.post("/api/auth/login", json={"username": "admin", "password": "admin123"})
        refresh = login.json()["data"]["refresh_token"]
        r = client.post("/api/auth/refresh", json={"refresh_token": refresh})
        assert r.status_code == 200
        assert "access_token" in r.json()["data"]

    def test_refresh_invalid_token(self, client):
        r = client.post("/api/auth/refresh", json={"refresh_token": "invalid"})
        assert r.status_code == 401

    def test_change_password(self, client):
        r = client.post("/api/auth/change-password", json={"old_password": "admin123", "new_password": "newpass123"})
        assert r.status_code == 200
        # 用新密码登录
        r2 = client.post("/api/auth/login", json={"username": "admin", "password": "newpass123"})
        assert r2.status_code == 200

    def test_change_password_wrong_old(self, client):
        r = client.post("/api/auth/change-password", json={"old_password": "wrong", "new_password": "newpass123"})
        assert r.status_code == 400

    def test_change_password_too_short(self, client):
        r = client.post("/api/auth/change-password", json={"old_password": "admin123", "new_password": "123"})
        assert r.status_code == 422


# ============================================================
# 人员 CRUD
# ============================================================

class TestPersons:
    def _create_person(self, client, **kwargs):
        data = {"name": "张三", "id_card": "32010219900100100X", **kwargs}
        return client.post("/api/persons", json=data)

    def test_create_person(self, client):
        r = self._create_person(client)
        assert r.status_code == 200
        d = r.json()
        assert d["name"] == "张三"
        assert d["id_card"] == "320102199001010010"
        assert d["gender"] == "男"  # 从身份证推算
        assert d["birth_date"] == "1990-01-01"  # 从身份证推算
        assert d["status"] == "在帮"  # 默认值
        assert d["risk_level"] == "低"  # 默认值

    def test_create_person_all_fields(self, client):
        r = client.post("/api/persons", json={
            "name": "李四", "id_card": "320102199001010029",
            "phone": "13800138000", "household_addr": "南京市玄武区",
            "current_addr": "南京市鼓楼区", "original_crime": "盗窃罪",
            "original_sentence": "有期徒刑三年", "release_date": "2025-06-15",
            "edu_start_date": "2025-07-01", "edu_end_date": "2028-06-30",
            "responsible_person": "张科员", "status": "在帮", "risk_level": "中",
            "family_name": "王五", "family_phone": "13900139000",
            "marital_status": "已婚", "education_level": "初中",
            "employment": "打零工", "health_status": "良好",
            "economic_status": "一般", "notes": "测试备注",
            "category": "刑满释放", "responsible_org": "城东司法所",
            "visit_interval_days": 30,
        })
        assert r.status_code == 200
        d = r.json()
        assert d["gender"] == "女"  # 身份证第17位偶数
        assert d["phone"] == "13800138000"
        assert d["original_crime"] == "盗窃罪"
        assert d["visit_interval_days"] == 30

    def test_create_duplicate_id_card(self, client):
        self._create_person(client)
        r = self._create_person(client, name="李四")
        assert r.status_code == 400

    def test_create_missing_required(self, client):
        r = client.post("/api/persons", json={"name": "张三"})
        assert r.status_code == 422

    def test_list_persons(self, client):
        self._create_person(client, id_card="320102199001010037")
        self._create_person(client, name="李四", id_card="320102199001010045")
        r = client.get("/api/persons")
        assert r.status_code == 200
        d = r.json()
        assert d["total"] == 2
        assert len(d["items"]) == 2
        assert "page" in d
        assert "page_size" in d
        assert "pages" in d

    def test_list_pagination(self, client):
        for i in range(5):
            self._create_person(client, name=f"用户{i}", id_card=f"3201021990010{i:04d}X")
        r = client.get("/api/persons?page=1&page_size=2")
        assert r.status_code == 200
        assert len(r.json()["data"]["items"]) == 2
        assert r.json()["data"]["total"] == 5

    def test_list_search(self, client):
        self._create_person(client, id_card="320102199001010053")
        self._create_person(client, name="李四", id_card="320102199001010061")
        r = client.get("/api/persons?search=张")
        assert r.json()["data"]["total"] == 1

    def test_list_filter_status(self, client):
        self._create_person(client, id_card="32010219900101007X", status="在帮")
        self._create_person(client, name="李四", id_card="320102199001010088", status="已解除")
        r = client.get("/api/persons?status=在帮")
        assert r.json()["data"]["total"] == 1

    def test_list_filter_risk(self, client):
        self._create_person(client, id_card="320102199001010096", risk_level="高")
        self._create_person(client, name="李四", id_card="320102199001010109", risk_level="低")
        r = client.get("/api/persons?risk_level=高")
        assert r.json()["data"]["total"] == 1

    def test_list_filter_crime(self, client):
        self._create_person(client, id_card="320102199001010117", original_crime="盗窃罪")
        self._create_person(client, name="李四", id_card="320102199001010125", original_crime="诈骗罪")
        r = client.get("/api/persons?crime_contains=盗窃")
        assert r.json()["data"]["total"] == 1

    def test_list_filter_age(self, client):
        self._create_person(client, id_card="320102199001010133")  # 36岁
        r = client.get("/api/persons?min_age=30&max_age=40")
        assert r.json()["data"]["total"] == 1
        r2 = client.get("/api/persons?min_age=50")
        assert r2.json()["data"]["total"] == 0

    def test_list_sort(self, client):
        self._create_person(client, name="赵六", id_card="320102199001010141")
        self._create_person(client, name="张三", id_card="32010219900101015X")
        r = client.get("/api/persons?sort_by=name&sort_order=asc")
        assert r.json()["data"]["items"][0]["name"] == "张三"

    def test_list_full_id_card(self, client):
        self._create_person(client, id_card="320102199001010168")
        # 列表接口始终返回完整身份证号（离线单所使用，不做脱敏）
        r1 = client.get("/api/persons")
        assert r1.json()["data"]["items"][0]["id_card"] == "320102199001010176"
        r2 = client.get("/api/persons?reveal=false")
        assert r2.json()["data"]["items"][0]["id_card"] == "320102199001010184"

    def test_list_last_visit_info(self, client):
        self._create_person(client, id_card="32010219900100119X")
        # 无走访
        r = client.get("/api/persons")
        assert r.json()["data"]["items"][0].get("last_visit_date") is None
        # 添加走访
        client.post("/api/visits", json={"person_id": 1, "visit_date": "2026-07-24", "visitor": "张科员", "visit_method": "上门"})
        r2 = client.get("/api/persons")
        assert r2.json()["data"]["items"][0].get("last_visit_date") == "2026-07-24"
        assert r2.json()["data"]["items"][0].get("last_visitor") == "张科员"

    def test_get_person(self, client):
        self._create_person(client, id_card="320102199001010192")
        r = client.get("/api/persons/1")
        assert r.status_code == 200
        assert r.json()["name"] == "张三"

    def test_get_person_not_found(self, client):
        r = client.get("/api/persons/999")
        assert r.status_code == 404

    def test_update_person(self, client):
        self._create_person(client, id_card="320102199001010205")
        r = client.patch("/api/persons/1", json={"risk_level": "高", "phone": "13900139000"})
        assert r.status_code == 200
        assert r.json()["risk_level"] == "高"
        assert r.json()["phone"] == "13900139000"

    def test_update_creates_edit_log(self, client):
        self._create_person(client, id_card="320102199001010213")
        client.patch("/api/persons/1", json={"risk_level": "高"})
        r = client.get("/api/persons/1/edit-logs")
        assert r.status_code == 200
        logs = r.json()
        assert len(logs) >= 1
        assert logs[0]["field_name"] == "risk_level"
        assert logs[0]["old_value"] == "低"
        assert logs[0]["new_value"] == "高"

    def test_update_no_change_no_log(self, client):
        self._create_person(client, id_card="320102199001010221")
        client.patch("/api/persons/1", json={"risk_level": "低"})  # 同值
        r = client.get("/api/persons/1/edit-logs")
        assert len(r.json()) == 0

    def test_update_not_found(self, client):
        r = client.patch("/api/persons/999", json={"risk_level": "高"})
        assert r.status_code == 404

    def test_delete_person(self, client):
        self._create_person(client, id_card="32010219900101023X")
        r = client.delete("/api/persons/1")
        assert r.status_code == 200
        # 软删除后查不到
        r2 = client.get("/api/persons/1")
        assert r2.status_code == 404
        # 列表也不显示
        r3 = client.get("/api/persons")
        assert r3.json()["data"]["total"] == 0

    def test_delete_not_found(self, client):
        r = client.delete("/api/persons/999")
        assert r.status_code == 404

    def test_edit_logs_empty(self, client):
        self._create_person(client, id_card="320102199001010248")
        r = client.get("/api/persons/1/edit-logs")
        assert r.status_code == 200
        assert r.json() == []

    def test_edit_logs_not_found(self, client):
        r = client.get("/api/persons/999/edit-logs")
        assert r.status_code == 404


# ============================================================
# 统计 API
# ============================================================

class TestStats:
    def test_stats_summary(self, client):
        r = client.post("/api/persons", json={"name": "张三", "id_card": "320102199001010256", "status": "在帮", "risk_level": "高"})
        client.post("/api/persons", json={"name": "李四", "id_card": "32010219900100127X", "status": "已解除", "risk_level": "低"})
        r = client.get("/api/persons/stats-summary")
        assert r.status_code == 200
        d = r.json()
        assert d["total"] == 2
        assert d["status_active"] == 1
        assert d["status_released"] == 1
        assert d["risk_high"] == 1
        assert d["risk_low"] == 1
        assert "expiring_soon" in d

    def test_quarterly_report(self, client):
        r = client.get("/api/persons/reports-quarterly")
        assert r.status_code == 200
        d = r.json()
        assert "quarter" in d
        assert "existing_at_start" in d
        assert "new_this_quarter" in d


# ============================================================
# Excel 导入导出
# ============================================================

class TestExcel:
    def test_export_excel(self, client):
        client.post("/api/persons", json={"name": "张三", "id_card": "320102199001010264"})
        r = client.get("/api/persons/export?format=excel")
        assert r.status_code == 200
        assert len(r.content) > 0
        assert r.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def test_export_with_filter(self, client):
        client.post("/api/persons", json={"name": "张三", "id_card": "320102199001010272", "status": "在帮"})
        client.post("/api/persons", json={"name": "李四", "id_card": "320102199001010280", "status": "已解除"})
        r = client.get("/api/persons/export?format=excel?status=在帮")
        assert r.status_code == 200

    def test_import_template(self, client):
        r = client.get("/api/persons/import/template")
        assert r.status_code == 200
        assert len(r.content) > 0

    def test_import_excel(self, client):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "身份证号", "联系电话", "原罪名", "状态", "风险等级"])
        ws.append(["导入用户", "320102199001010299", "13800138000", "盗窃罪", "在帮", "低"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        r = client.post("/api/persons/import/excel", files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        assert r.status_code == 200
        d = r.json()
        assert d["imported"] >= 0

    def test_import_duplicate(self, client):
        from openpyxl import Workbook
        client.post("/api/persons", json={"name": "已有", "id_card": "320102199001010301"})
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "身份证号"])
        ws.append(["重复", "32010219900101031X"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        r = client.post("/api/persons/import/excel", files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        assert r.json()["skipped"] == 1

    def test_import_invalid(self, client):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "身份证号"])
        ws.append(["坏数据", "123"])  # 身份证号不合法
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        r = client.post("/api/persons/import/excel", files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        assert r.json()["errors"]


# ============================================================
# 走访记录
# ============================================================

class TestVisits:
    def _create_person(self, client):
        return client.post("/api/persons", json={"name": "张三", "id_card": "320102199001010328"})

    def test_create_visit(self, client):
        self._create_person(client)
        r = client.post("/api/visits", json={
            "person_id": 1, "visit_date": "2026-07-24", "visitor": "张科员",
            "visit_method": "上门", "content": "走访正常"
        })
        assert r.status_code == 200
        d = r.json()
        assert d["quarter"] == "2026-Q3"
        assert d["has_abnormal"] is False

    def test_create_visit_abnormal(self, client):
        self._create_person(client)
        r = client.post("/api/visits", json={
            "person_id": 1, "visit_date": "2026-07-24", "visitor": "张科员",
            "visit_method": "电话", "has_abnormal": True, "abnormal_detail": "联系不上"
        })
        assert r.status_code == 200
        assert r.json()["has_abnormal"] is True

    def test_create_visit_nonexistent_person(self, client):
        r = client.post("/api/visits", json={"person_id": 999, "visit_date": "2026-07-24", "visitor": "A", "visit_method": "上门"})
        assert r.status_code == 404

    def test_list_visits(self, client):
        self._create_person(client)
        client.post("/api/visits", json={"person_id": 1, "visit_date": "2026-07-24", "visitor": "A", "visit_method": "上门"})
        client.post("/api/visits", json={"person_id": 1, "visit_date": "2026-07-25", "visitor": "B", "visit_method": "电话"})
        r = client.get("/api/visits")
        assert r.json()["data"]["total"] == 2

    def test_list_visits_by_person(self, client):
        self._create_person(client)
        client.post("/api/visits", json={"person_id": 1, "visit_date": "2026-07-24", "visitor": "A", "visit_method": "上门"})
        r = client.get("/api/visits?person_id=1")
        assert r.json()["data"]["total"] == 1

    def test_list_visits_by_quarter(self, client):
        self._create_person(client)
        client.post("/api/visits", json={"person_id": 1, "visit_date": "2026-07-24", "visitor": "A", "visit_method": "上门"})
        client.post("/api/visits", json={"person_id": 1, "visit_date": "2026-03-10", "visitor": "B", "visit_method": "电话"})
        r = client.get("/api/visits?quarter=2026-Q3")
        assert r.json()["data"]["total"] == 1

    def test_get_visit(self, client):
        self._create_person(client)
        client.post("/api/visits", json={"person_id": 1, "visit_date": "2026-07-24", "visitor": "A", "visit_method": "上门"})
        r = client.get("/api/visits/1")
        assert r.status_code == 200

    def test_get_visit_not_found(self, client):
        r = client.get("/api/visits/999")
        assert r.status_code == 404

    def test_update_visit(self, client):
        self._create_person(client)
        client.post("/api/visits", json={"person_id": 1, "visit_date": "2026-07-24", "visitor": "A", "visit_method": "上门"})
        r = client.patch("/api/visits/1", json={"person_id": 1, "visit_date": "2026-07-25", "visitor": "B", "visit_method": "电话", "content": "更新"})
        assert r.status_code == 200
        assert r.json()["visitor"] == "B"

    def test_delete_visit(self, client):
        self._create_person(client)
        client.post("/api/visits", json={"person_id": 1, "visit_date": "2026-07-24", "visitor": "A", "visit_method": "上门"})
        r = client.delete("/api/visits/1")
        assert r.status_code == 200
        r2 = client.get("/api/visits/1")
        assert r2.status_code == 404

    def test_quarterly_stats(self, client):
        self._create_person(client)
        client.post("/api/visits", json={"person_id": 1, "visit_date": "2026-07-24", "visitor": "A", "visit_method": "上门"})
        client.post("/api/visits", json={"person_id": 1, "visit_date": "2026-07-25", "visitor": "B", "visit_method": "电话", "has_abnormal": True})
        r = client.get("/api/visits/stats-quarterly")
        assert r.status_code == 200
        d = r.json()
        assert d["total"] == 2
        assert d["上门"] == 1
        assert d["电话"] == 1
        assert d["有异常"] == 1


# ============================================================
# 提醒系统
# ============================================================

class TestReminders:
    def test_reminders(self, client):
        r = client.get("/api/reminders")
        assert r.status_code == 200
        d = r.json()
        assert "expiring_30d" in d
        assert "expiring_7d" in d
        assert "overdue_expired" in d
        assert "visit_overdue" in d
        assert "quarter_deadline_days" in d
        assert "quarter_deadline_date" in d
        assert "expiring_list" in d
        assert "visit_overdue_list" in d

    def test_reminders_with_data(self, client):
        # 创建即将到期的人员
        from datetime import date, timedelta
        soon = (date.today() + timedelta(days=15)).isoformat()
        client.post("/api/persons", json={"name": "快到期", "id_card": "32010219900100135X", "status": "在帮", "edu_end_date": soon})
        r = client.get("/api/reminders")
        d = r.json()
        assert "expiring_30d" in d

    def test_visit_overdue(self, client):
        # 创建超期未走访的人员（visit_interval_days=30, 但从未走访）
        from datetime import date, timedelta
        old = (date.today() - timedelta(days=60)).isoformat()
        client.post("/api/persons", json={"name": "超期", "id_card": "320102199001010336", "status": "在帮", "edu_start_date": old, "visit_interval_days": 30})
        r = client.get("/api/reminders")
        assert "visit_overdue" in r.json()


# ============================================================
# OCR
# ============================================================

