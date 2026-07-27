"""人员 CRUD API 测试"""
import pytest
from fastapi.testclient import TestClient
from sqlmodel import SQLModel, Session, create_engine
from sqlmodel.pool import StaticPool

from app.main import app
from app.core.database import get_session
from app.models import User
from app.core.security import hash_password


@pytest.fixture
def client():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    SQLModel.metadata.create_all(engine)

    with Session(engine) as session:
        user = User(
            username="testadmin",
            hashed_password=hash_password("test123"),
            real_name="测试管理员",
            role="director",
        )
        session.add(user)
        session.commit()

    def get_test_session():
        with Session(engine) as session:
            yield session

    app.dependency_overrides[get_session] = get_test_session

    with TestClient(app) as c:
        # 登录获取token
        login_resp = c.post("/api/auth/login", json={
            "username": "testadmin",
            "password": "test123",
        })
        token = login_resp.json()["access_token"]
        c.headers["Authorization"] = f"Bearer {token}"
        yield c

    app.dependency_overrides.clear()


def test_create_person(client):
    """新增人员"""
    resp = client.post("/api/persons", json={
        "name": "张三",
        "id_card": "320102199001011234",
        "phone": "13800138000",
        "original_crime": "盗窃罪",
        "status": "在帮",
    })
    assert resp.status_code == 201
    data = resp.json()
    assert data["name"] == "张三"
    assert data["gender"] == "男"  # 从身份证推算
    assert data["birth_date"] == "1990-01-01"  # 从身份证推算


def test_create_person_duplicate_id_card(client):
    """身份证号重复"""
    client.post("/api/persons", json={
        "name": "张三",
        "id_card": "320102199001011234",
    })
    resp = client.post("/api/persons", json={
        "name": "李四",
        "id_card": "320102199001011234",
    })
    assert resp.status_code == 400


def test_list_persons(client):
    """列表查询"""
    # 先创建两个
    client.post("/api/persons", json={"name": "张三", "id_card": "320102199001011234"})
    client.post("/api/persons", json={"name": "李四", "id_card": "320102199002021235"})

    resp = client.get("/api/persons")
    assert resp.status_code == 200
    data = resp.json()
    assert data["total"] == 2
    assert len(data["items"]) == 2


def test_search_persons(client):
    """搜索"""
    client.post("/api/persons", json={"name": "张三", "id_card": "320102199001011234"})
    client.post("/api/persons", json={"name": "李四", "id_card": "320102199002021235"})

    resp = client.get("/api/persons?search=张")
    assert resp.status_code == 200
    assert resp.json()["total"] == 1
    assert resp.json()["items"][0]["name"] == "张三"


def test_get_person_detail(client):
    """详情"""
    create_resp = client.post("/api/persons", json={
        "name": "张三",
        "id_card": "320102199001011234",
    })
    person_id = create_resp.json()["id"]

    resp = client.get(f"/api/persons/{person_id}")
    assert resp.status_code == 200
    assert resp.json()["name"] == "张三"


def test_update_person(client):
    """修改"""
    create_resp = client.post("/api/persons", json={
        "name": "张三",
        "id_card": "320102199001011234",
    })
    person_id = create_resp.json()["id"]

    resp = client.put(f"/api/persons/{person_id}", json={
        "phone": "13900139000",
        "status": "已解除",
    })
    assert resp.status_code == 200
    assert resp.json()["phone"] == "13900139000"
    assert resp.json()["status"] == "已解除"


def test_delete_person(client):
    """软删除"""
    create_resp = client.post("/api/persons", json={
        "name": "张三",
        "id_card": "320102199001011234",
    })
    person_id = create_resp.json()["id"]

    resp = client.delete(f"/api/persons/{person_id}")
    assert resp.status_code == 200

    # 删除后查不到
    resp = client.get(f"/api/persons/{person_id}")
    assert resp.status_code == 404


def test_filter_by_status(client):
    """按状态筛选"""
    client.post("/api/persons", json={"name": "张三", "id_card": "320102199001011234", "status": "在帮"})
    client.post("/api/persons", json={"name": "李四", "id_card": "320102199002021235", "status": "已解除"})

    resp = client.get("/api/persons?status=在帮")
    assert resp.json()["total"] == 1
    assert resp.json()["items"][0]["name"] == "张三"


# ========== 修改历史 API ==========

def test_get_edit_logs(client):
    """获取修改历史"""
    # 创建人员
    create_resp = client.post("/api/persons", json={
        "name": "张三",
        "id_card": "320102199001011234",
        "phone": "13800138000",
    })
    person_id = create_resp.json()["id"]

    # 修改两次
    client.put(f"/api/persons/{person_id}", json={"phone": "13900139000"})
    client.put(f"/api/persons/{person_id}", json={"status": "已解除"})

    # 获取修改历史
    resp = client.get(f"/api/persons/{person_id}/edit-logs")
    assert resp.status_code == 200
    logs = resp.json()
    assert len(logs) == 2
    # 按时间倒序，最新的在前
    assert logs[0]["field_name"] == "status"
    assert logs[0]["new_value"] == "已解除"
    assert logs[1]["field_name"] == "phone"
    assert logs[1]["new_value"] == "13900139000"
    assert logs[1]["old_value"] == "13800138000"


def test_get_edit_logs_not_found(client):
    """人员不存在时返回404"""
    resp = client.get("/api/persons/999/edit-logs")
    assert resp.status_code == 404


def test_get_edit_logs_empty(client):
    """无修改记录时返回空列表"""
    create_resp = client.post("/api/persons", json={
        "name": "张三",
        "id_card": "320102199001011234",
    })
    person_id = create_resp.json()["id"]

    resp = client.get(f"/api/persons/{person_id}/edit-logs")
    assert resp.status_code == 200
    assert resp.json() == []


# ========== 统计汇总 API ==========

def test_stats_summary(client):
    """统计汇总"""
    # 创建不同状态的人员
    client.post("/api/persons", json={
        "name": "张三", "id_card": "320102199001011234",
        "status": "在帮", "risk_level": "高",
    })
    client.post("/api/persons", json={
        "name": "李四", "id_card": "320102199002021235",
        "status": "在帮", "risk_level": "低",
    })
    client.post("/api/persons", json={
        "name": "王五", "id_card": "320102199003031236",
        "status": "已解除", "risk_level": "中",
    })
    client.post("/api/persons", json={
        "name": "赵六", "id_card": "320102199004041237",
        "status": "脱管", "risk_level": "低",
    })

    resp = client.get("/api/persons/stats/summary")
    assert resp.status_code == 200
    data = resp.json()
    assert data["total"] == 4
    assert data["在帮"] == 2
    assert data["已解除"] == 1
    assert data["脱管"] == 1
    assert data["重点关注"] == 0
    assert data["risk_high"] == 1
    assert data["risk_medium"] == 1
    assert data["risk_low"] == 2
    assert data["expiring_soon"] == 0  # 未设置截止日期


def test_stats_summary_expiring_soon(client):
    """即将到期统计"""
    from datetime import date, timedelta
    end_date = (date.today() + timedelta(days=30)).isoformat()
    client.post("/api/persons", json={
        "name": "张三", "id_card": "320102199001011234",
        "status": "在帮", "edu_end_date": end_date,
    })
    # 已解除的不应计入即将到期
    client.post("/api/persons", json={
        "name": "李四", "id_card": "320102199002021235",
        "status": "已解除", "edu_end_date": end_date,
    })

    resp = client.get("/api/persons/stats/summary")
    data = resp.json()
    assert data["expiring_soon"] == 1


# ========== Excel 导出 API ==========

def test_export_excel(client):
    """导出 Excel"""
    client.post("/api/persons", json={
        "name": "张三", "id_card": "320102199001011234",
        "status": "在帮", "risk_level": "高",
    })
    client.post("/api/persons", json={
        "name": "李四", "id_card": "320102199002021235",
        "status": "已解除", "risk_level": "低",
    })

    resp = client.get("/api/persons/export/excel")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    # 文件名包含日期
    assert "xlsx" in resp.headers["content-disposition"]

    # 验证 Excel 内容
    from openpyxl import load_workbook
    wb = load_workbook(filename=__import__("io").BytesIO(resp.content))
    ws = wb.active
    assert ws.max_row == 3  # 表头 + 2行数据
    assert ws.cell(1, 1).value == "姓名"
    # 按 updated_at 倒序，第二个创建的“李四”排在前面
    names = {ws.cell(r, 1).value for r in range(2, 4)}
    assert names == {"张三", "李四"}


def test_export_excel_with_filter(client):
    """导出 Excel 带筛选"""
    client.post("/api/persons", json={
        "name": "张三", "id_card": "320102199001011234", "status": "在帮",
    })
    client.post("/api/persons", json={
        "name": "李四", "id_card": "320102199002021235", "status": "已解除",
    })

    resp = client.get("/api/persons/export/excel?status=在帮")
    assert resp.status_code == 200
    from openpyxl import load_workbook
    wb = load_workbook(filename=__import__("io").BytesIO(resp.content))
    ws = wb.active
    assert ws.max_row == 2  # 表头 + 1行数据
    assert ws.cell(2, 1).value == "张三"


# ========== Excel 批量导入 API ==========

IMPORT_HEADERS = [
    "姓名", "身份证号", "性别", "出生日期", "户籍地址", "现住址",
    "联系电话", "原罪名", "原判刑期", "释放日期", "帮教起始日期", "帮教截止日期",
    "帮教责任人", "状态", "风险等级", "家属姓名", "家属电话", "婚姻状况",
    "文化程度", "就业情况", "身体状况", "经济状况", "备注", "责任单位",
]

def _make_test_id_card(body17: str) -> str:
    """根据前17位生成合法的18位身份证号（含校验位）"""
    weights = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2]
    check_chars = "10X98765432"
    total = sum(int(body17[i]) * weights[i] for i in range(17))
    return body17 + check_chars[total % 11]

def _make_import_excel(rows: list[list]) -> bytes:
    """生成导入用的 Excel 文件（表头 + 数据行）"""
    from openpyxl import Workbook
    import io as _io
    wb = Workbook()
    ws = wb.active
    ws.append(IMPORT_HEADERS)
    for row in rows:
        ws.append(row)
    stream = _io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()


def test_import_template(client):
    """下载导入模板，验证返回 xlsx"""
    resp = client.get("/api/persons/import/template")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "xlsx" in resp.headers["content-disposition"]

    from openpyxl import load_workbook
    import io
    wb = load_workbook(filename=io.BytesIO(resp.content))
    ws = wb.active
    # 3行：表头、示例、说明
    assert ws.max_row == 3
    # 表头第一列是"姓名"
    assert ws.cell(1, 1).value == "姓名"
    # 表头最后一列是"责任单位"
    assert ws.cell(1, ws.max_column).value == "责任单位"
    # 示例数据
    assert ws.cell(2, 1).value == "张三"
    # 说明行
    assert "必填" in str(ws.cell(3, 1).value)


def test_import_excel_success(client):
    """上传有效 Excel，验证导入成功"""
    id1 = _make_test_id_card("11010119900307123")  # 校验位 8
    id2 = _make_test_id_card("11010119881225123")  # 校验位 2

    excel_bytes = _make_import_excel([
        ["王五", id1, "", "", None, None, None, None, None, None, None, None, None, "在帮", "低"],
        ["赵六", id2, "", "", None, None, None, None, None, None, None, None, None, "已解除", "中"],
    ])

    resp = client.post(
        "/api/persons/import/excel",
        files={"file": ("test.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["success"] is True
    assert data["total_rows"] == 2
    assert data["imported"] == 2
    assert data["skipped"] == 0
    assert len(data["errors"]) == 0

    # 验证数据已入库
    resp = client.get("/api/persons?search=王五")
    assert resp.json()["total"] == 1
    assert resp.json()["items"][0]["gender"] == "男"  # 从身份证推算


def test_import_excel_duplicate(client):
    """上传含重复身份证的 Excel，验证跳过"""
    id1 = _make_test_id_card("11010119900307123")

    # 先在数据库中创建一条
    client.post("/api/persons", json={"name": "已有人员", "id_card": id1})

    # 上传包含同一身份证号的 Excel
    excel_bytes = _make_import_excel([
        ["新人员A", id1, "", "", None, None, None, None, None, None, None, None, None, "在帮", "低"],
    ])

    resp = client.post(
        "/api/persons/import/excel",
        files={"file": ("test.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["total_rows"] == 1
    assert data["imported"] == 0
    assert data["skipped"] == 1
    assert len(data["errors"]) == 1
    assert data["errors"][0]["field"] == "id_card"
    assert "已存在" in data["errors"][0]["message"]


def test_import_excel_invalid(client):
    """上传格式错误的 Excel，验证错误报告"""
    bad_id = "320102199001011234"  # 校验位不正确
    id_valid = _make_test_id_card("11010119900307123")

    excel_bytes = _make_import_excel([
        ["无身份证", "", "", "", None, None, None, None, None, None, None, None, None, "在帮", "低"],  # 缺少必填
        ["错误证件", bad_id, "", "", None, None, None, None, None, None, None, None, None, "在帮", "低"],  # 身份证格式错
        ["正常人", id_valid, "", "", None, None, None, None, None, None, None, None, None, "无效状态", "低"],  # 无效枚举
    ])

    resp = client.post(
        "/api/persons/import/excel",
        files={"file": ("test.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["total_rows"] == 3
    assert data["imported"] == 0
    assert data["skipped"] == 3
    assert len(data["errors"]) >= 3

    # 检查错误类型
    error_fields = {e["field"] for e in data["errors"]}
    assert "name" in error_fields or "id_card" in error_fields  # 缺少必填
    assert "id_card" in error_fields  # 格式错误
    assert "status" in error_fields  # 无效枚举


def test_import_excel_mixed(client):
    """混合场景：部分成功部分失败"""
    id_valid = _make_test_id_card("11010119900307123")
    bad_id = "320102199001011234"

    excel_bytes = _make_import_excel([
        ["正常人", id_valid, "", "", None, None, None, None, None, None, None, None, None, "在帮", "低"],
        ["错误人", bad_id, "", "", None, None, None, None, None, None, None, None, None, "在帮", "低"],
    ])

    resp = client.post(
        "/api/persons/import/excel",
        files={"file": ("test.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["total_rows"] == 2
    assert data["imported"] == 1
    assert data["skipped"] == 1
    assert data["success"] is True  # 至少导入了一条
